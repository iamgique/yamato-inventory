import pymysql
import openpyxl


class yamato:
    uid = 1
    filename = './excel/Inventory_20150416.xlsx'
    worksheet = None
    sheetname = 'Inventory for check stock'
    colsmap = {
        'date': 'A',
        'warehouse_code': 'B',
        'item_code': 'C',
        'item_name': 'D',
        'item_color': 'E',
        'location_no': 'F',
        'qty_on_hand': 'G'
    }

    @classmethod
    def load_workbook(cls):
        workbook = openpyxl.load_workbook(cls.filename)
        cls.worksheet = workbook[cls.sheetname]

    @classmethod
    def get_items(cls):
        #print cls.worksheet['A1'].value
        for row, item in enumerate(cls.worksheet.rows, start=1):
            print cls.worksheet[cls.colsmap['qty_on_hand'] + str(row)].value

    @classmethod
    def get_mat_from_sap_matcode(cls, sap_matcode=''):
        pass


    @classmethod
    def create_items(cls, items=[]):
        pass

    @classmethod
    def gen_uid(cls, current=1):
        return current + 1


yamato.load_workbook()
yamato.get_items()

#print ws['A10'].value

#print len(ws.rows)
#print len(ws.columns)


    #print row.cell('A').value
#    print row()
    #for cell in row:
    #    print(cell.value)
